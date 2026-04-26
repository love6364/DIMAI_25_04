import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import io

# ---------------- PAGE CONFIG ----------------
st.set_page_config(page_title="Add Price Columns", page_icon="💎", layout="centered")

st.title("💎 Add Updated Price & Difference")

st.write("Upload your Excel file → Download updated file with formula")

# ---------------- FILE UPLOAD ----------------
file = st.file_uploader("Upload Excel File", type=["xlsx"])

# ---------------- FUNCTION ----------------
def process_file(file):
    wb = load_workbook(file)
    ws = wb.active

    # Detect Cost / Cts column
    headers = [cell.value for cell in ws[1]]
    cost_col = None

    for i, h in enumerate(headers):
        if h and "cost" in str(h).lower() and "cts" in str(h).lower():
            cost_col = i + 1
            break

    if cost_col is None:
        raise Exception("❌ 'Cost / Cts.' column not found")

    # Add new columns
    last_col = ws.max_column
    upd_col = last_col + 1
    diff_col = last_col + 2

    ws.cell(1, upd_col).value = "Updated Price"
    ws.cell(1, diff_col).value = "Difference"

    # Column letters
    cost_letter = get_column_letter(cost_col)
    upd_letter = get_column_letter(upd_col)

    # Apply formula
    for row in range(2, ws.max_row + 1):
        formula = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
        ws.cell(row, diff_col).value = formula

    # Save file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# ---------------- PROCESS ----------------
if file:
    try:
        output = process_file(file)

        st.success("✅ Columns added successfully!")

        st.download_button(
            label="📥 Download Updated File",
            data=output,
            file_name="updated_price_file.xlsx"
        )

    except Exception as e:
        st.error(str(e))